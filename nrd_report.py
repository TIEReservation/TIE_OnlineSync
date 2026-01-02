"""
nrd_report.py - Night Report Dashboard for TIE Hotels & Resorts (Streamlit Integration)
Updated to display properties as columns and metrics as rows
"""

import streamlit as st
from datetime import date, timedelta
import pandas as pd
from typing import Dict, List, Any, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import io
import calendar

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ============================================================================
# CONFIGURATION (matching inventory.py)
# ============================================================================

PROPERTY_SHORT_NAMES = {
    "Le Poshe Beach view": "LPBV",
    "La Millionaire Resort": "LMR",
    "Le Poshe Luxury": "LePL",
    "Le Poshe Suite": "LPS",
    "La Paradise Residency": "LaPR",
    "La Paradise Luxury": "LaPL",
    "La Villa Heritage": "LVH",
    "Le Pondy Beachside": "LPB",
    "Le Royce Villa": "LRV",
    "La Tamara Luxury": "LTL",
    "La Antilia Luxury": "LAL",
    "La Tamara Suite": "LTS",
    "Le Park Resort": "LePR",
    "Villa Shakti": "VS",
    "Eden Beach Resort": "EBR",
    "Le Terra": "LT",
    "La Coromandel Luxury": "LCL",
    "Happymates Forest Retreat": "HFR"
}

# Full inventory from inventory.py
PROPERTY_INVENTORY = {
    "Le Poshe Beach view": {"all": ["101","102","201","202","203","204","301","302","303","304","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203","204"]},
    "La Millionaire Resort": {"all": ["101","102","103","105","201","202","203","204","205","206","207","208","301","302","303","304","305","306","307","308","401","402","Day Use 1","Day Use 2","Day Use 3","Day Use 4","Day Use 5","No Show"],"three_bedroom":["203","204","205"]},
    "Le Poshe Luxury": {"all": ["101","102","201","202","203","204","205","301","302","303","304","305","401","402","403","404","405","501","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203","204","205"]},
    "Le Poshe Suite": {"all": ["601","602","603","604","701","702","703","704","801","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "La Paradise Residency": {"all": ["101","102","103","201","202","203","301","302","303","304","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203"]},
    "La Paradise Luxury": {"all": ["101","102","103","201","202","203","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203"]},
    "La Villa Heritage": {"all": ["101","102","103","201","202","203","301","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203"]},
    "Le Pondy Beachside": {"all": ["101","102","201","202","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "Le Royce Villa": {"all": ["101","102","201","202","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "La Tamara Luxury": {"all": ["101","102","103","104","105","106","201","202","203","204","205","206","301","302","303","304","305","306","401","402","403","404","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203","204","205","206"]},
    "La Antilia Luxury": {"all": ["101","201","202","203","204","301","302","303","304","401","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203","204"]},
    "La Tamara Suite": {"all": ["101","102","103","104","201","202","203","204","205","206","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203","204","205","206"]},
    "Le Park Resort": {"all": ["111","222","333","444","555","666","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "Villa Shakti": {"all": ["101","102","201","201A","202","203","301","301A","302","303","401","Day Use 1","Day Use 2","No Show"],"three_bedroom":["203"]},
    "Eden Beach Resort": {"all": ["101","102","103","201","202","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "Le Terra": {"all": ["101","102","103","104","105","106","107","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "La Coromandel Luxury": {"all": ["101","102","103","201","202","203","204","205","206","301","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]},
    "Happymates Forest Retreat": {"all": ["101","102","Day Use 1","Day Use 2","No Show"],"three_bedroom":[]}  
}

property_mapping = {
    "La Millionaire Luxury Resort": "La Millionaire Resort",
    "Le Poshe Beach View": "Le Poshe Beach view",
    "Le Poshe Beach view": "Le Poshe Beach view",
    "Le Poshe Beach VIEW": "Le Poshe Beach view",
    "Le Poshe Beachview": "Le Poshe Beach view",
    "Millionaire": "La Millionaire Resort",
    "Le Pondy Beach Side": "Le Pondy Beachside",
    "Le Teera": "Le Terra"
}

reverse_mapping = {c: [] for c in set(property_mapping.values())}
for v, c in property_mapping.items():
    reverse_mapping[c].append(v)

mob_mapping = {
    "Booking": ["BOOKING"],
    "Direct": ["Direct"],
    "Bkg-Direct": ["Bkg-Direct"],
    "Agoda": ["Agoda"],
    "Go-MMT": ["Goibibo", "MMT", "Go-MMT", "MAKEMYTRIP"],
    "Walk-In": ["Walk-In"],
    "TIE Group": ["TIE Group"],
    "Stayflexi": ["STAYFLEXI_GHA"],
    "Airbnb": ["Airbnb"],
    "Social Media": ["Social Media"],
    "Expedia": ["Expedia"],
    "Cleartrip": ["Cleartrip"],
    "Website": ["Stayflexi Booking Engine"],
}

# ============================================================================
# HELPER FUNCTIONS (from inventory.py)
# ============================================================================

def normalize_property(name: str) -> str:
    return property_mapping.get(name.strip(), name.strip())

def sanitize_string(v: Any, default: str = "") -> str:
    return str(v).strip() if v is not None else default

def safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(float(v)) if v not in [None, "", " "] else default
    except:
        return default

def safe_float(v: Any, default: float = 0) -> float:
    try:
        return float(v) if v not in [None, "", " "] else default
    except:
        return default

# ============================================================================
# DATA NORMALIZATION (from inventory.py)
# ============================================================================

def normalize_booking(row: Dict, is_online: bool) -> Optional[Dict]:
    """Normalize booking data - EXACT copy from inventory.py"""
    try:
        bid = sanitize_string(row.get("booking_id") or row.get("id"))
        status_field = "booking_status" if is_online else "plan_status"
        status = sanitize_string(row.get(status_field, "")).title()
        if status not in ["Confirmed", "Completed"]: return None
        pay = sanitize_string(row.get("payment_status")).title()
        if pay not in ["Fully Paid", "Partially Paid"]: return None
        ci = date.fromisoformat(row["check_in"])
        co = date.fromisoformat(row["check_out"])
        if co <= ci: return None
        days_field = "room_nights" if is_online else "no_of_days"
        days = safe_int(row.get(days_field)) or (co - ci).days
        if days <= 0: days = 1
        p = normalize_property(row.get("property_name") if not is_online else row.get("property"))

        if is_online:
            total_amount = safe_float(row.get("booking_amount")) or 0.0
            gst = safe_float(row.get("gst")) or 0.0
            tax = safe_float(row.get("ota_tax")) or 0.0
            commission = safe_float(row.get("ota_commission")) or 0.0
            room_charges = total_amount - gst - tax
        else:
            total_amount = safe_float(row.get("total_tariff")) or 0.0
            gst = tax = commission = 0.0
            room_charges = total_amount

        receivable = total_amount - gst - tax - commission
        if receivable < 0: receivable = 0.0

        identifier = row.get("id") if is_online else row.get("booking_id")
        identifier_str = str(identifier) if identifier is not None else ""

        return {
            "type": "online" if is_online else "direct",
            "property": p,
            "booking_id": bid,
            "guest_name": sanitize_string(row.get("guest_name")),
            "mobile_no": sanitize_string(row.get("guest_phone") if is_online else row.get("mobile_no")),
            "total_pax": safe_int(row.get("total_pax")),
            "check_in": str(ci),
            "check_out": str(co),
            "days": days,
            "room_no": sanitize_string(row.get("room_no")).title(),
            "mob": sanitize_string(row.get("mode_of_booking") if is_online else row.get("mob")),
            "plan": sanitize_string(row.get("rate_plans") if is_online else row.get("breakfast")),
            "room_charges": room_charges,
            "gst": gst,
            "tax": tax,
            "total_amount": total_amount,
            "commission": commission,
            "receivable": receivable,
            "advance": safe_float(row.get("total_payment_made") if is_online else row.get("advance_amount")),
            "advance_mop": sanitize_string(row.get("advance_mop")),
            "balance": safe_float(row.get("balance_due") if is_online else row.get("balance_amount")),
            "balance_mop": sanitize_string(row.get("balance_mop")),
            "booking_status": status,
            "payment_status": pay,
            "submitted_by": sanitize_string(row.get("submitted_by")),
            "modified_by": sanitize_string(row.get("modified_by")),
            "remarks": sanitize_string(row.get("remarks")),
            "db_id": identifier_str,
        }
    except Exception as e:
        logging.warning(f"normalize failed: {e}")
        return None

# ============================================================================
# DATA FETCHING (from inventory.py)
# ============================================================================

def load_combined_bookings(supabase, property: str, start_date: date, end_date: date) -> List[Dict]:
    """EXACT copy from inventory.py"""
    prop = normalize_property(property)
    query_props = [prop] + reverse_mapping.get(prop, [])
    combined: List[Dict] = []

    try:
        q = supabase.table("reservations").select("*").in_("property_name", query_props).lte("check_in", str(end_date)).gte("check_out", str(start_date)).in_("plan_status", ["Confirmed", "Completed"]).in_("payment_status", ["Partially Paid", "Fully Paid"]).execute()
        for r in q.data or []:
            norm = normalize_booking(r, is_online=False)
            if norm: combined.append(norm)
    except Exception as e:
        logging.error(f"Direct query error: {e}")

    try:
        q = supabase.table("online_reservations").select("*").in_("property", query_props).lte("check_in", str(end_date)).gte("check_out", str(start_date)).in_("booking_status", ["Confirmed", "Completed"]).in_("payment_status", ["Partially Paid", "Fully Paid"]).execute()
        for r in q.data or []:
            norm = normalize_booking(r, is_online=True)
            if norm: combined.append(norm)
    except Exception as e:
        logging.error(f"Online query error: {e}")

    return combined

# ============================================================================
# FILTERING & ASSIGNMENT (from inventory.py)
# ============================================================================

def filter_bookings_for_day(bookings: List[Dict], day: date) -> List[Dict]:
    """EXACT copy from inventory.py"""
    return [b.copy() for b in bookings if date.fromisoformat(b["check_in"]) <= day < date.fromisoformat(b["check_out"])]

def assign_inventory_numbers(daily_bookings: List[Dict], property: str):
    """EXACT copy from inventory.py"""
    assigned, over = [], []
    inv = PROPERTY_INVENTORY.get(property, {"all": []})["all"]
    inv_lookup = {i.strip().lower(): i for i in inv}
    
    room_bookings = {}
    sorted_bookings = sorted(daily_bookings, key=lambda x: (x.get("check_in", ""), x.get("booking_id", "")))

    for b in sorted_bookings:
        raw_room = str(b.get("room_no", "") or "").strip()
        booking_id = b.get("booking_id", "Unknown")
        
        if not raw_room:
            over.append(b)
            continue

        requested = [r.strip() for r in raw_room.split(",") if r.strip()]
        assigned_rooms = []
        is_overbooking = False

        for r in requested:
            key = r.lower()
            if key not in inv_lookup:
                is_overbooking = True
                break
            room_name = inv_lookup[key]
            
            if room_name in room_bookings and room_bookings[room_name] != booking_id:
                is_overbooking = True
                break
            assigned_rooms.append(room_name)

        if is_overbooking or not assigned_rooms:
            over.append(b)
            continue

        for room in assigned_rooms:
            room_bookings[room] = booking_id

        days = max(b.get("days", 1), 1)
        receivable = b.get("receivable", 0.0)
        num_rooms = len(assigned_rooms)
        total_nights = days * num_rooms
        per_night = receivable / total_nights if total_nights > 0 else 0.0
        base_pax = b["total_pax"] // num_rooms if num_rooms else 0
        rem = b["total_pax"] % num_rooms if num_rooms else 0

        for idx, room in enumerate(assigned_rooms):
            nb = b.copy()
            nb["assigned_room"] = room
            nb["room_no"] = room
            nb["total_pax"] = base_pax + (1 if idx < rem else 0)
            nb["per_night"] = per_night
            nb["is_primary"] = (idx == 0)
            assigned.append(nb)

    return assigned, over

# ============================================================================
# STATISTICS EXTRACTION (from inventory.py)
# ============================================================================

def extract_stats_from_assigned(assigned: List[Dict], target_date: date, mob_types: List[str]) -> Dict:
    """Extract stats matching inventory.py logic"""
    
    def to_float(val):
        try:
            return float(val) if val not in [None, "", " "] else 0.0
        except:
            return 0.0
    
    def to_int(val):
        try:
            return int(val) if val not in [None, "", " "] else 0
        except:
            return 0
    
    dtd = {m: {"rooms":0,"value":0.0,"comm":0.0,"gst":0.0,"tax":0.0,"pax":0} for m in mob_types}
    dtd_rooms = 0
    dtd_value = 0.0
    dtd_comm = 0.0
    dtd_gst = 0.0
    dtd_tax = 0.0
    dtd_pax = 0

    for booking in assigned:
        check_in_date = date.fromisoformat(booking["check_in"])
        is_check_in_day = (target_date == check_in_date)
        is_primary = booking.get("is_primary", False)
        
        # Count rooms
        dtd_rooms += 1
        
        # Per night value
        per_night = to_float(booking.get("per_night", 0))
        dtd_value += per_night
        
        # Only count full amounts on check-in day for primary booking
        if is_check_in_day and is_primary:
            dtd_comm += to_float(booking.get("commission", 0))
            dtd_gst += to_float(booking.get("gst", 0))
            dtd_tax += to_float(booking.get("tax", 0))
        
        # Pax
        dtd_pax += to_int(booking.get("total_pax", 0))
        
        # MOB breakdown
        mob_raw = sanitize_string(booking.get("mob", ""))
        mob = next((m for m, vs in mob_mapping.items() if mob_raw.upper() in [v.upper() for v in vs]), "Booking")
        
        dtd[mob]["rooms"] += 1
        dtd[mob]["value"] += per_night
        dtd[mob]["pax"] += to_int(booking.get("total_pax", 0))
        
        if is_check_in_day and is_primary:
            dtd[mob]["comm"] += to_float(booking.get("commission", 0))
            dtd[mob]["gst"] += to_float(booking.get("gst", 0))
            dtd[mob]["tax"] += to_float(booking.get("tax", 0))
    
    # Calculate ARR for each MOB
    for m in mob_types:
        r = dtd[m]["rooms"]
        dtd[m]["arr"] = dtd[m]["value"] / r if r > 0 else 0.0

    dtd["Total"] = {
        "rooms": dtd_rooms,
        "value": dtd_value,
        "arr": dtd_value / dtd_rooms if dtd_rooms > 0 else 0.0,
        "comm": dtd_comm,
        "gst": dtd_gst,
        "tax": dtd_tax,
        "pax": dtd_pax
    }

    return dtd

# ============================================================================
# EXCEL EXPORT - NEW FORMAT
# ============================================================================

def export_multiple_days_to_excel(all_dates_data: List[Dict], year: int, month: int) -> bytes:
    """Export multiple days to Excel with separate sheet for each date"""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Styles
    header_fill = PatternFill(start_color="00B8A9", end_color="00B8A9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=12, color="FFFFFF")
    date_font = Font(bold=True, size=11, color="FFFFFF")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get sorted property list
    all_props = sorted(PROPERTY_SHORT_NAMES.keys(), key=lambda x: PROPERTY_SHORT_NAMES.get(x, x))
    short_names = [PROPERTY_SHORT_NAMES[p] for p in all_props]
    
    # Create a sheet for each date
    for day_data in all_dates_data:
        target_date = day_data["date"]
        metrics = day_data["metrics"]
        totals = day_data["totals"]
        
        # Create sheet with date as name
        sheet_name = target_date.strftime("%d-%b")
        ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1
        
        # Title row
        ws.merge_cells(f'A{current_row}:{get_column_letter(len(short_names) + 3)}{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="TIE Hotels & Resorts")
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = header_fill
        
        # Date in last column
        date_cell = ws.cell(row=current_row, column=len(short_names) + 4)
        date_cell.value = target_date.strftime("%d-%b-%y")
        date_cell.font = date_font
        date_cell.fill = header_fill
        date_cell.alignment = center_align
        current_row += 1
        
        # Subtitle row
        ws.merge_cells(f'A{current_row}:{get_column_letter(len(short_names) + 4)}{current_row}')
        subtitle_cell = ws.cell(row=current_row, column=1, value="Overall Report for the day")
        subtitle_cell.alignment = center_align
        subtitle_cell.fill = white_fill
        current_row += 1
        
        # Header row with property names
        header_row = current_row
        ws.cell(row=header_row, column=1, value="")
        
        col_idx = 2
        for short_name in short_names:
            cell = ws.cell(row=header_row, column=col_idx, value=short_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            col_idx += 1
        
        # D.T.D and M.T.D columns
        dtd_cell = ws.cell(row=header_row, column=col_idx, value="D.T.D")
        dtd_cell.font = header_font
        dtd_cell.fill = header_fill
        dtd_cell.alignment = center_align
        dtd_cell.border = thin_border
        col_idx += 1
        
        mtd_cell = ws.cell(row=header_row, column=col_idx, value="M.T.D")
        mtd_cell.font = header_font
        mtd_cell.fill = header_fill
        mtd_cell.alignment = center_align
        mtd_cell.border = thin_border
        
        current_row += 1
        
        # Data rows
        row_labels = ["Rooms Available", "Rooms Sold", "Occ %", "GST", "Commission", 
                     "Receivable", "Receivable Per Night", "ARR"]
        
        for label in row_labels:
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.border = thin_border
            
            col_idx = 2
            for prop in all_props:
                m = metrics[prop]
                
                if label == "Rooms Available":
                    value = m["rooms_available"]
                elif label == "Rooms Sold":
                    value = m["rooms_sold"]
                elif label == "Occ %":
                    value = f"{m['occupancy']:.0f}%"
                elif label == "GST":
                    value = int(m["gst"])
                elif label == "Commission":
                    value = int(m["commission"])
                elif label == "Receivable":
                    value = int(m["receivable"])
                elif label == "Receivable Per Night":
                    value = int(m["receivable_per_night"])
                elif label == "ARR":
                    value = int(m["arr"])
                
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = thin_border
                
                if label == "Occ %":
                    cell.alignment = center_align
                elif isinstance(value, (int, float)):
                    cell.alignment = right_align
                    if value > 0:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = center_align
                
                col_idx += 1
            
            # D.T.D column
            if label == "Rooms Available":
                dtd_val = totals["rooms_available"]
            elif label == "Rooms Sold":
                dtd_val = totals["rooms_sold"]
            elif label == "Occ %":
                dtd_val = f"{totals['occupancy']:.0f}%"
            elif label == "GST":
                dtd_val = int(totals["gst"])
            elif label == "Commission":
                dtd_val = int(totals["commission"])
            elif label == "Receivable":
                dtd_val = int(totals["receivable"])
            elif label == "Receivable Per Night":
                dtd_val = int(totals["receivable_per_night"])
            elif label == "ARR":
                dtd_val = int(totals["arr"])
            
            cell = ws.cell(row=current_row, column=col_idx, value=dtd_val)
            cell.border = thin_border
            if label == "Occ %":
                cell.alignment = center_align
            elif isinstance(dtd_val, (int, float)):
                cell.alignment = right_align
                if dtd_val > 0:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = center_align
            col_idx += 1
            
            # M.T.D column (same as D.T.D for now)
            cell = ws.cell(row=current_row, column=col_idx, value=dtd_val)
            cell.border = thin_border
            if label == "Occ %":
                cell.alignment = center_align
            elif isinstance(dtd_val, (int, float)):
                cell.alignment = right_align
                if dtd_val > 0:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = center_align
            
            current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 22
        for i in range(2, len(short_names) + 5):
            ws.column_dimensions[get_column_letter(i)].width = 10
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ============================================================================
# STREAMLIT UI
# ============================================================================

@st.cache_data(ttl=300)
def load_month_bookings(property: str, year: int, month: int):
    """Cache bookings for entire month"""
    supabase = st.session_state.get('supabase_client')
    if supabase is None:
        from supabase import create_client
        try:
            supabase_url = st.secrets["supabase"]["url"]
            supabase_key = st.secrets["supabase"]["key"]
            supabase = create_client(supabase_url, supabase_key)
        except:
            supabase_url = os.getenv("SUPABASE_URL", "https://oxbrezracnmazucnnqox.supabase.co")
            supabase_key = os.getenv("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im94YnJlenJhY25tYXp1Y25ucW94Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NTM3NjUxMTgsImV4cCI6MjA2OTM0MTExOH0.nqBK2ZxntesLY9qYClpoFPVnXOW10KrzF-UI_DKjbKo")
            supabase = create_client(supabase_url, supabase_key)
    
    start = date(year, month, 1)
    num_days = calendar.monthrange(year, month)[1]
    end = date(year, month, num_days)
    
    return load_combined_bookings(supabase, property, start, end)

def create_daily_summary_table(metrics: Dict, totals: Dict, mtd_totals: Dict = None) -> pd.DataFrame:
    """Create table with metrics as rows and properties as columns"""
    
    # Get sorted property list
    all_props = sorted(PROPERTY_SHORT_NAMES.keys(), key=lambda x: PROPERTY_SHORT_NAMES.get(x, x))
    short_names = [PROPERTY_SHORT_NAMES[p] for p in all_props]
    
    # Build the table data
    rows = []
    
    row_labels = ["Rooms Available", "Rooms Sold", "Occ %", "GST", "Commission", 
                 "Receivable", "Receivable Per Night", "ARR"]
    
    for label in row_labels:
        row = {"Metric": label}
        
        # Add each property's value
        for prop in all_props:
            short_name = PROPERTY_SHORT_NAMES[prop]
            m = metrics[prop]
            
            if label == "Rooms Available":
                value = m["rooms_available"]
            elif label == "Rooms Sold":
                value = m["rooms_sold"]
            elif label == "Occ %":
                value = f"{m['occupancy']:.0f}%"
            elif label == "GST":
                value = f"{int(m['gst']):,}" if m['gst'] > 0 else 0
            elif label == "Commission":
                value = f"{int(m['commission']):,}" if m['commission'] > 0 else 0
            elif label == "Receivable":
                value = f"{int(m['receivable']):,}" if m['receivable'] > 0 else 0
            elif label == "Receivable Per Night":
                value = f"{int(m['receivable_per_night']):,}" if m['receivable_per_night'] > 0 else 0
            elif label == "ARR":
                value = f"{int(m['arr']):,}" if m['arr'] > 0 else 0
            
            row[short_name] = value
        
        # Add D.T.D total
        if label == "Rooms Available":
            dtd_val = totals["rooms_available"]
        elif label == "Rooms Sold":
            dtd_val = totals["rooms_sold"]
        elif label == "Occ %":
            dtd_val = f"{totals['occupancy']:.0f}%"
        elif label == "GST":
            dtd_val = f"{int(totals['gst']):,}" if totals['gst'] > 0 else 0
        elif label == "Commission":
            dtd_val = f"{int(totals['commission']):,}" if totals['commission'] > 0 else 0
        elif label == "Receivable":
            dtd_val = f"{int(totals['receivable']):,}" if totals['receivable'] > 0 else 0
        elif label == "Receivable Per Night":
            dtd_val = f"{int(totals['receivable_per_night']):,}" if totals['receivable_per_night'] > 0 else 0
        elif label == "ARR":
            dtd_val = f"{int(totals['arr']):,}" if totals['arr'] > 0 else 0
        
        row["D.T.D"] = dtd_val
        
        # Add M.T.D total (if provided)
        if mtd_totals:
            if label == "Rooms Available":
                mtd_val = mtd_totals["rooms_available"]
            elif label == "Rooms Sold":
                mtd_val = mtd_totals["rooms_sold"]
            elif label == "Occ %":
                mtd_val = f"{mtd_totals['occupancy']:.0f}%"
            elif label == "GST":
                mtd_val = f"{int(mtd_totals['gst']):,}" if mtd_totals['gst'] > 0 else 0
            elif label == "Commission":
                mtd_val = f"{int(mtd_totals['commission']):,}" if mtd_totals['commission'] > 0 else 0
            elif label == "Receivable":
                mtd_val = f"{int(mtd_totals['receivable']):,}" if mtd_totals['receivable'] > 0 else 0
            elif label == "Receivable Per Night":
                mtd_val = f"{int(mtd_totals['receivable_per_night']):,}" if mtd_totals['receivable_per_night'] > 0 else 0
            elif label == "ARR":
                mtd_val = f"{int(mtd_totals['arr']):,}" if mtd_totals['arr'] > 0 else 0
            
            row["M.T.D"] = mtd_val
        
        rows.append(row)
    
    # Create DataFrame
    columns = ["Metric"] + short_names + ["D.T.D"]
    if mtd_totals:
        columns.append("M.T.D")
    
    df = pd.DataFrame(rows, columns=columns)
    return df

def highlight_nrd_table(df):
    """Apply styling to NRD table"""
    def style_row(row):
        styles = [''] * len(row)
        # Bold the first column (Metric names)
        styles[0] = 'font-weight: bold'
        
        # Highlight D.T.D and M.T.D columns
        if 'D.T.D' in df.columns:
            dtd_idx = df.columns.get_loc('D.T.D')
            styles[dtd_idx] = 'background-color: #D3D3D3; font-weight: bold'
        
        if 'M.T.D' in df.columns:
            mtd_idx = df.columns.get_loc('M.T.D')
            styles[mtd_idx] = 'background-color: #D3D3D3; font-weight: bold'
        
        return styles
    
    return df.style.apply(style_row, axis=1)

def show_nrd_report():
    """Display the Night Report Dashboard in Streamlit"""
    st.title("📊 Night Report Dashboard (NRD)")
    st.markdown("**TIE Hotels & Resorts - Overall Daily Report**")
    
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()
    
    # Month and Year selectors
    today = date.today()
    col1, col2 = st.columns(2)
    
    with col1:
        year = st.selectbox("Year", list(range(today.year-5, today.year+6)), index=5, key="nrd_year")
    
    with col2:
        month = st.selectbox("Month", list(range(1, 13)), index=today.month-1, 
                            format_func=lambda x: calendar.month_name[x], key="nrd_month")
    
    st.markdown("---")
    
    # Generate dates for the month
    num_days = calendar.monthrange(year, month)[1]
    month_dates = [date(year, month, d) for d in range(1, num_days + 1)]
    
    # Only show dates up to today
    month_dates = [d for d in month_dates if d <= today]
    
    if not month_dates:
        st.warning(f"No data available for {calendar.month_name[month]} {year}")
        return
    
    mob_types = list(mob_mapping.keys())
    
    # Process all dates in the month
    with st.spinner(f"Loading data for {calendar.month_name[month]} {year}..."):
        all_dates_data = []
        
        # Calculate M.T.D totals
        mtd_totals = {
            "rooms_available": 0,
            "rooms_sold": 0,
            "gst": 0.0,
            "commission": 0.0,
            "receivable": 0.0,
            "receivable_per_night": 0.0,
        }
        
        for selected_date in month_dates:
            # Pre-load all bookings for all properties for this date
            date_metrics = {}
            
            for prop in PROPERTY_SHORT_NAMES.keys():
                # Get inventory count
                inv_data = PROPERTY_INVENTORY.get(prop, {"all": []})
                all_rooms = inv_data["all"]
                total_inventory = len([i for i in all_rooms if not i.startswith(("Day Use", "No Show"))])
                
                # Load bookings
                bookings = load_month_bookings(prop, year, month)
                
                # Filter bookings for this day
                daily = filter_bookings_for_day(bookings, selected_date)
                
                if not daily:
                    date_metrics[prop] = {
                        "rooms_available": total_inventory,
                        "rooms_sold": 0,
                        "occupancy": 0.0,
                        "gst": 0.0,
                        "commission": 0.0,
                        "receivable": 0.0,
                        "receivable_per_night": 0.0,
                        "arr": 0.0
                    }
                    continue
                
                # Assign inventory
                assigned, over = assign_inventory_numbers(daily, prop)
                
                # Extract stats
                stats = extract_stats_from_assigned(assigned, selected_date, mob_types)
                total_stats = stats["Total"]
                
                rooms_sold = total_stats["rooms"]
                occupancy = (rooms_sold / total_inventory * 100) if total_inventory > 0 else 0.0
                
                date_metrics[prop] = {
                    "rooms_available": total_inventory,
                    "rooms_sold": rooms_sold,
                    "occupancy": occupancy,
                    "gst": total_stats["gst"],
                    "commission": total_stats["comm"],
                    "receivable": total_stats["value"],
                    "receivable_per_night": total_stats["value"],
                    "arr": total_stats["arr"]
                }
            
            # Calculate D.T.D totals for this date
            dtd_totals = {
                "rooms_available": sum(m["rooms_available"] for m in date_metrics.values()),
                "rooms_sold": sum(m["rooms_sold"] for m in date_metrics.values()),
                "gst": sum(m["gst"] for m in date_metrics.values()),
                "commission": sum(m["commission"] for m in date_metrics.values()),
                "receivable": sum(m["receivable"] for m in date_metrics.values()),
                "receivable_per_night": sum(m["receivable_per_night"] for m in date_metrics.values()),
            }
            dtd_totals["occupancy"] = (
                dtd_totals["rooms_sold"] / dtd_totals["rooms_available"] * 100
                if dtd_totals["rooms_available"] > 0 else 0.0
            )
            dtd_totals["arr"] = (
                dtd_totals["receivable_per_night"] / dtd_totals["rooms_sold"]
                if dtd_totals["rooms_sold"] > 0 else 0.0
            )
            
            # Update M.T.D running totals
            mtd_totals["rooms_available"] += dtd_totals["rooms_available"]
            mtd_totals["rooms_sold"] += dtd_totals["rooms_sold"]
            mtd_totals["gst"] += dtd_totals["gst"]
            mtd_totals["commission"] += dtd_totals["commission"]
            mtd_totals["receivable"] += dtd_totals["receivable"]
            mtd_totals["receivable_per_night"] += dtd_totals["receivable_per_night"]
            
            # Calculate M.T.D percentages
            current_mtd = {
                "rooms_available": mtd_totals["rooms_available"],
                "rooms_sold": mtd_totals["rooms_sold"],
                "gst": mtd_totals["gst"],
                "commission": mtd_totals["commission"],
                "receivable": mtd_totals["receivable"],
                "receivable_per_night": mtd_totals["receivable_per_night"],
            }
            current_mtd["occupancy"] = (
                current_mtd["rooms_sold"] / current_mtd["rooms_available"] * 100
                if current_mtd["rooms_available"] > 0 else 0.0
            )
            current_mtd["arr"] = (
                current_mtd["receivable_per_night"] / current_mtd["rooms_sold"]
                if current_mtd["rooms_sold"] > 0 else 0.0
            )
            
            # Store data for this date
            all_dates_data.append({
                "date": selected_date,
                "metrics": date_metrics,
                "dtd_totals": dtd_totals,
                "mtd_totals": current_mtd.copy()
            })
    
    st.success(f"✅ Data loaded for {len(month_dates)} days in {calendar.month_name[month]} {year}")
    
    # Display each date's report
    for day_data in all_dates_data:
        selected_date = day_data["date"]
        date_metrics = day_data["metrics"]
        dtd_totals = day_data["dtd_totals"]
        mtd_totals = day_data["mtd_totals"]
        
        # Display header
        st.markdown(f"### 🏨 TIE Hotels & Resorts - {selected_date.strftime('%d %b %Y (%A)')}")
        st.markdown("**Overall Report for the day**")
        
        # Create and display the summary table
        summary_df = create_daily_summary_table(date_metrics, dtd_totals, mtd_totals)
        
        # Apply styling
        styled_df = highlight_nrd_table(summary_df)
        
        # Display the table
        st.dataframe(styled_df, use_container_width=True, height=400, hide_index=True)
        
        # Display summary metrics in columns
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "D.T.D Rooms Sold", 
                f"{dtd_totals['rooms_sold']:,}",
                f"{dtd_totals['occupancy']:.1f}% Occupancy"
            )
        
        with col2:
            st.metric(
                "D.T.D Revenue", 
                f"₹{dtd_totals['receivable']:,.0f}",
                f"₹{dtd_totals['arr']:,.0f} ARR"
            )
        
        with col3:
            st.metric(
                "M.T.D Rooms Sold", 
                f"{mtd_totals['rooms_sold']:,}",
                f"{mtd_totals['occupancy']:.1f}% Occupancy"
            )
        
        with col4:
            st.metric(
                "M.T.D Revenue", 
                f"₹{mtd_totals['receivable']:,.0f}",
                f"₹{mtd_totals['arr']:,.0f} ARR"
            )
        
        st.markdown("---")
    
    # Download buttons at the end for all dates
    st.subheader("📥 Download Reports")
    col1, col2 = st.columns(2)
    
    with col1:
        # CSV export - combine all dates
        all_csv_rows = []
        for day_data in all_dates_data:
            summary_df = create_daily_summary_table(
                day_data["metrics"], 
                day_data["dtd_totals"], 
                day_data["mtd_totals"]
            )
            summary_df.insert(0, "Date", day_data["date"].strftime("%d-%b-%Y"))
            all_csv_rows.append(summary_df)
        
        combined_csv_df = pd.concat(all_csv_rows, ignore_index=True)
        csv = combined_csv_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download Complete CSV Report",
            data=csv,
            file_name=f"TIE_NRD_Report_{year}_{month:02d}_Complete.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col2:
        # Excel export - all dates in separate sheets
        excel_export_data = [{
            "date": d["date"],
            "metrics": d["metrics"],
            "totals": d["dtd_totals"]
        } for d in all_dates_data]
        
        excel_data = export_multiple_days_to_excel(excel_export_data, year, month)
        st.download_button(
            label="📥 Download Complete Excel Report",
            data=excel_data,
            file_name=f"TIE_NRD_Report_{year}_{month:02d}_Complete.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

if __name__ == "__main__":
    show_nrd_report()
